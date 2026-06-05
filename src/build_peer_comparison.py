"""
Stage 10: Five-Company Semiconductor Peer Comparison Model

Reads validated company-level financials and metrics CSVs.
Produces:
  - data/processed/semiconductor_peer_metrics.csv
  - output/semiconductor_peer_comparison.xlsx (9 tabs)

Does NOT modify any company-level files.
Does NOT rerun extraction pipelines or read raw JSON.
"""

import os
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── paths ────────────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROCESSED_DIR = os.path.join(BASE_DIR, "data", "processed")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
MANUAL_DIR = os.path.join(BASE_DIR, "data", "manual_checks")

CSV_OUT = os.path.join(PROCESSED_DIR, "semiconductor_peer_metrics.csv")
XLSX_OUT = os.path.join(OUTPUT_DIR, "semiconductor_peer_comparison.xlsx")

# ── company metadata ─────────────────────────────────────────────────────────

COMPANIES = [
    {"name": "Qualcomm Incorporated", "ticker": "QCOM", "short_id": "qualcomm",
     "fy_end": "Sep (~Sep 29)", "model": "Fabless + Licensing"},
    {"name": "Advanced Micro Devices Inc.", "ticker": "AMD", "short_id": "amd",
     "fy_end": "Dec (~Dec 28)", "model": "Fabless"},
    {"name": "Nvidia Corp.", "ticker": "NVDA", "short_id": "nvidia",
     "fy_end": "Jan (~Jan 26)", "model": "Fabless"},
    {"name": "Intel Corp.", "ticker": "INTC", "short_id": "intel",
     "fy_end": "Dec (~Dec 28)", "model": "IDM (Integrated Device Manufacturer)"},
    {"name": "Broadcom Inc.", "ticker": "AVGO", "short_id": "broadcom",
     "fy_end": "Nov (~Nov 3)", "model": "Fabless + Infrastructure Software"},
]

TICKER_MAP = {c["name"]: c["ticker"] for c in COMPANIES}
SHORT_MAP = {c["name"]: c["short_id"] for c in COMPANIES}
MODEL_MAP = {c["ticker"]: c["model"] for c in COMPANIES}
FYEND_MAP = {c["ticker"]: c["fy_end"] for c in COMPANIES}

TICKER_ORDER = ["QCOM", "AMD", "NVDA", "INTC", "AVGO"]

# ── company-specific limitation notes ────────────────────────────────────────

COMPANY_NOTES = {
    "QCOM": "Licensing revenue (QTL segment) is high-margin IP licensing, inflating consolidated margins relative to pure-chip peers. FY2025 Q4 Net Income = -$3.1B likely reflects a one-time charge.",
    "AMD": "Total Debt FY2023 Q1 missing (XBRL tag coverage begins Q2 after Xilinx acquisition accounting settled).",
    "NVDA": "FY numbering offset +1 from calendar year (FY2026 = Feb 2025 - Jan 2026). CapEx FY2024 Q1-Q3 sourced from 8-K press releases, pending manual confirmation; FCF/TTM FCF empty for those quarters. Revenue scale ($130B+ TTM) dramatically larger than peers.",
    "INTC": "IDM model: operates own fabs, CapEx $14-26B/year (10-20x fabless peers), structurally negative FCF. Cash tag includes restricted cash (~3% / $447M overstatement vs peers). Cash FY2025 Q2/Q3 missing. FY2024 includes ~$15.9B non-recurring impairment/restructuring charges distorting margins.",
    "AVGO": "FY2024 is 53-week VMware transition year: +$16B acquired revenue, ~$8-10B acquired intangible amortization depressing margins. Total Debt uses gross principal (~$2B / 3% above carrying value vs peers). Revenue mix ~50/50 semiconductors + infrastructure software post-VMware.",
}

# ── load data ────────────────────────────────────────────────────────────────

def load_all_financials():
    """Load and concatenate all company financials CSVs."""
    frames = []
    for pattern in glob.glob(os.path.join(PROCESSED_DIR, "*_financials_*.csv")):
        df = pd.read_csv(pattern)
        df["ticker"] = df["company"].map(TICKER_MAP)
        frames.append(df)
    combined = pd.concat(frames, ignore_index=True)
    return combined


def load_all_metrics():
    """Load and concatenate all company metrics CSVs."""
    frames = []
    for pattern in glob.glob(os.path.join(PROCESSED_DIR, "*_metrics.csv")):
        df = pd.read_csv(pattern)
        df["ticker"] = df["company"].map(TICKER_MAP)
        frames.append(df)
    combined = pd.concat(frames, ignore_index=True)
    return combined


def load_missing_metrics():
    """Load all missing-metrics files from manual_checks."""
    frames = []
    for path in glob.glob(os.path.join(MANUAL_DIR, "*_missing_metrics_*.csv")):
        df = pd.read_csv(path)
        filename = os.path.basename(path)
        short_id = filename.split("_missing_metrics")[0]
        ticker_lookup = {c["short_id"]: c["ticker"] for c in COMPANIES}
        df["ticker"] = ticker_lookup.get(short_id, short_id)
        company_lookup = {c["short_id"]: c["name"] for c in COMPANIES}
        df["company"] = company_lookup.get(short_id, short_id)
        frames.append(df)
    if frames:
        return pd.concat(frames, ignore_index=True)
    return pd.DataFrame()


# ── build consolidated CSV ───────────────────────────────────────────────────

def build_consolidated_csv(fin_df, met_df):
    """Merge financials and metrics into one long-format CSV."""

    # financials columns
    fin_cols = fin_df[["company", "ticker", "fiscal_year", "fiscal_quarter",
                       "metric_name", "value", "unit", "extraction_method",
                       "source_reference", "derived_from",
                       "requires_manual_review"]].copy()
    fin_cols["formula"] = ""

    # metrics columns
    met_cols = met_df[["company", "ticker", "fiscal_year", "fiscal_quarter",
                       "metric_name", "value", "unit"]].copy()
    met_cols["extraction_method"] = "calculated"
    met_cols["source_reference"] = ""
    met_cols["derived_from"] = ""
    met_cols["requires_manual_review"] = ""
    met_cols["formula"] = met_df["formula"]

    combined = pd.concat([fin_cols, met_cols], ignore_index=True)

    # sort: ticker order, then fiscal_year desc, quarter order
    q_order = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
    combined["_ticker_rank"] = combined["ticker"].map(
        {t: i for i, t in enumerate(TICKER_ORDER)})
    combined["_q_rank"] = combined["fiscal_quarter"].map(q_order)
    combined.sort_values(["_ticker_rank", "fiscal_year", "_q_rank",
                          "metric_name"],
                         ascending=[True, False, True, True], inplace=True)
    combined.drop(columns=["_ticker_rank", "_q_rank"], inplace=True)
    combined.reset_index(drop=True, inplace=True)

    combined.to_csv(CSV_OUT, index=False)
    print(f"Wrote {len(combined)} rows to {CSV_OUT}")
    return combined


# ── Excel helpers ────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(bold=True, size=13)
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
NOTE_FONT = Font(italic=True, size=10, color="666666")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
MISSING_FILL = PatternFill("solid", fgColor="F2DCDB")
DERIVED_FILL = PatternFill("solid", fgColor="FFFFCC")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)

USD_FMT = '#,##0'
USD_M_FMT = '#,##0.0'
PCT_FMT = '0.0%'
GROWTH_FMT = '0.0%'


def write_header_row(ws, row, headers, widths=None):
    """Write a styled header row and freeze the row below it."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    if widths:
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w


def write_title(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT


def write_note(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = NOTE_FONT
    cell.alignment = Alignment(wrap_text=True)


def fmt_usd_millions(val):
    """Convert raw USD to millions for display."""
    if pd.isna(val) or val == "" or val is None:
        return None
    try:
        return float(val) / 1_000_000
    except (ValueError, TypeError):
        return None


def get_period_label(row):
    return f"FY{row['fiscal_year']} {row['fiscal_quarter']}"


# ── Tab builders ─────────────────────────────────────────────────────────────

def get_latest_quarter(fin_df, ticker):
    """Find the latest quarter with data for a given ticker."""
    t_df = fin_df[fin_df["ticker"] == ticker].copy()
    t_df = t_df[t_df["metric_name"] == "Revenue"]
    t_df = t_df.dropna(subset=["value"])
    t_df = t_df[t_df["value"] != ""]
    if t_df.empty:
        return None, None
    q_order = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
    t_df["_q"] = t_df["fiscal_quarter"].map(q_order)
    t_df = t_df.sort_values(["fiscal_year", "_q"], ascending=[False, False])
    latest = t_df.iloc[0]
    return int(latest["fiscal_year"]), latest["fiscal_quarter"]


def get_metric_value(df, ticker, fy, fq, metric):
    """Get a single metric value."""
    mask = ((df["ticker"] == ticker) &
            (df["fiscal_year"] == fy) &
            (df["fiscal_quarter"] == fq) &
            (df["metric_name"] == metric))
    rows = df[mask]
    if rows.empty:
        return None
    val = rows.iloc[0]["value"]
    if pd.isna(val) or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def build_peer_summary(wb, fin_df, met_df):
    """Tab 1: Peer Summary — latest quarter + TTM by company."""
    ws = wb.create_sheet("Peer Summary")

    write_title(ws, 1, "Semiconductor Peer Comparison — Summary")
    write_note(ws, 2, 1,
               "Peer comparability is imperfect: fiscal year-ends differ, "
               "business models vary (IDM vs fabless vs licensing), "
               "and accounting tags are not fully standardized across companies. "
               "See Validation tab for details.")

    # find latest quarter per company
    latest_periods = {}
    for ticker in TICKER_ORDER:
        fy, fq = get_latest_quarter(fin_df, ticker)
        latest_periods[ticker] = (fy, fq)

    # summary metrics to show
    summary_metrics = [
        ("Revenue", "USD", "fin", True),
        ("Gross Profit", "USD", "fin", True),
        ("Operating Income", "USD", "fin", True),
        ("Net Income", "USD", "fin", True),
        ("R&D Expense", "USD", "fin", True),
        ("Capital Expenditure", "USD", "fin", True),
        ("Operating Cash Flow", "USD", "fin", True),
        ("Free Cash Flow", "USD", "met", True),
        ("Cash and Cash Equivalents", "USD", "fin", True),
        ("Total Debt", "USD", "fin", True),
        ("Net Cash (Debt)", "USD", "met", True),
        ("Gross Margin", "%", "met", False),
        ("Operating Margin", "%", "met", False),
        ("Free-Cash-Flow Margin", "%", "met", False),
        ("R&D as % of Revenue", "%", "met", False),
        ("YoY Revenue Growth", "%", "met", False),
        ("TTM Revenue", "USD", "met", True),
        ("TTM Operating Income", "USD", "met", True),
        ("TTM Free Cash Flow", "USD", "met", True),
    ]

    # header row
    row = 4
    headers = ["Metric"] + [f"{t}\n{FYEND_MAP[t]}" for t in TICKER_ORDER]
    widths = [28] + [20] * 5
    write_header_row(ws, row, headers, widths)

    # period label row
    row = 5
    ws.cell(row=row, column=1, value="Latest Quarter").font = SECTION_FONT
    for ci, ticker in enumerate(TICKER_ORDER, 2):
        fy, fq = latest_periods[ticker]
        ws.cell(row=row, column=ci,
                value=f"FY{fy} {fq}").font = Font(bold=True)
        ws.cell(row=row, column=ci).alignment = Alignment(horizontal="center")

    # data rows
    row = 6
    for metric_name, unit, source, is_usd in summary_metrics:
        ws.cell(row=row, column=1, value=metric_name)
        df_src = fin_df if source == "fin" else met_df
        for ci, ticker in enumerate(TICKER_ORDER, 2):
            fy, fq = latest_periods[ticker]
            val = get_metric_value(df_src, ticker, fy, fq, metric_name)
            cell = ws.cell(row=row, column=ci)
            if val is None:
                cell.value = "N/A"
                cell.fill = MISSING_FILL
                cell.alignment = Alignment(horizontal="center")
            elif is_usd:
                cell.value = val / 1_000_000
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = val
                cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal="right")
        row += 1

    # units note
    row += 1
    write_note(ws, row, 1, "USD values in millions. Margins and growth rates as decimals (0.50 = 50%).")

    # business model annotations
    row += 2
    ws.cell(row=row, column=1, value="Business Model Notes").font = SECTION_FONT
    row += 1
    for ticker in TICKER_ORDER:
        ws.cell(row=row, column=1, value=ticker).font = Font(bold=True)
        ws.cell(row=row, column=2, value=MODEL_MAP[ticker])
        row += 1

    # company-specific notes
    row += 1
    ws.cell(row=row, column=1,
            value="Company-Specific Comparability Notes").font = SECTION_FONT
    row += 1
    for ticker in TICKER_ORDER:
        ws.cell(row=row, column=1, value=ticker).font = Font(bold=True)
        note_cell = ws.cell(row=row, column=2, value=COMPANY_NOTES[ticker])
        note_cell.alignment = Alignment(wrap_text=True)
        note_cell.font = NOTE_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1


def build_quarterly_trends(wb, fin_df, met_df):
    """Tab 2: Quarterly Trends — all metrics, all companies, all quarters."""
    ws = wb.create_sheet("Quarterly Trends")
    write_title(ws, 1, "Quarterly Trends — All Metrics")
    write_note(ws, 2, 1, "USD values in millions. Periods are company-specific fiscal quarters.")

    combined = pd.concat([
        fin_df[["ticker", "fiscal_year", "fiscal_quarter", "metric_name",
                "value", "unit", "extraction_method"]],
        met_df[["ticker", "fiscal_year", "fiscal_quarter", "metric_name",
                "value", "unit"]].assign(extraction_method="calculated"),
    ], ignore_index=True)

    q_order = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
    combined["_q"] = combined["fiscal_quarter"].map(q_order)
    combined.sort_values(["metric_name", "ticker", "fiscal_year", "_q"],
                         inplace=True)

    headers = ["Metric", "Ticker", "Period", "Value", "Unit", "Extraction Method"]
    widths = [28, 8, 14, 18, 8, 22]
    write_header_row(ws, 4, headers, widths)

    row = 5
    for _, r in combined.iterrows():
        ws.cell(row=row, column=1, value=r["metric_name"])
        ws.cell(row=row, column=2, value=r["ticker"])
        ws.cell(row=row, column=3,
                value=f"FY{r['fiscal_year']} {r['fiscal_quarter']}")

        val = r["value"]
        cell = ws.cell(row=row, column=4)
        if pd.isna(val) or val == "":
            cell.value = "N/A"
            cell.fill = MISSING_FILL
        else:
            try:
                fval = float(val)
                if r["unit"] == "USD":
                    cell.value = fval / 1_000_000
                    cell.number_format = '#,##0'
                elif r["unit"] == "%":
                    cell.value = fval
                    cell.number_format = '0.0%'
                else:
                    cell.value = fval
                    cell.number_format = '#,##0.00'
            except (ValueError, TypeError):
                cell.value = val

        ws.cell(row=row, column=5, value=r["unit"])

        em = r.get("extraction_method", "")
        ws.cell(row=row, column=6, value=em if pd.notna(em) else "")
        if em in ("derived_ytd_difference", "fiscal_year_end_balance"):
            ws.cell(row=row, column=4).fill = DERIVED_FILL
        elif em == "missing_requires_review":
            ws.cell(row=row, column=4).fill = MISSING_FILL

        row += 1


def _build_pivot_tab(wb, met_df, fin_df, tab_name, title, metrics_list, source_map):
    """
    Generic builder for tabs 3-6.
    metrics_list: [(metric_name, unit_type, source)]
    source_map: 'fin' or 'met'
    Pivots to companies-as-columns layout.
    """
    ws = wb.create_sheet(tab_name)
    write_title(ws, 1, title)
    write_note(ws, 2, 1,
               "USD values in millions. Periods are company-specific fiscal quarters. "
               "Yellow = derived value. Pink = missing / requires review.")

    # build all periods across all companies
    all_periods = set()
    for df_src in [fin_df, met_df]:
        for _, r in df_src.iterrows():
            all_periods.add((r["ticker"], int(r["fiscal_year"]),
                             r["fiscal_quarter"]))

    # sort periods per ticker
    q_order = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}

    for metric_idx, (metric_name, unit_type, source) in enumerate(metrics_list):
        df_src = fin_df if source == "fin" else met_df

        # section header
        start_row = 4 + metric_idx * 16
        ws.cell(row=start_row, column=1,
                value=metric_name).font = SECTION_FONT

        # header
        headers = ["Period"] + TICKER_ORDER
        widths = [14] + [18] * 5
        write_header_row(ws, start_row + 1, headers, widths)
        ws.freeze_panes = None  # we'll set freeze on first tab only

        # collect all (fy, fq) combinations and sort descending
        periods = set()
        for ticker in TICKER_ORDER:
            t_data = df_src[(df_src["ticker"] == ticker) &
                            (df_src["metric_name"] == metric_name)]
            for _, r in t_data.iterrows():
                periods.add((int(r["fiscal_year"]), r["fiscal_quarter"]))

        periods = sorted(periods, key=lambda x: (x[0], q_order.get(x[1], 0)),
                         reverse=True)

        data_row = start_row + 2
        for fy, fq in periods:
            ws.cell(row=data_row, column=1, value=f"FY{fy} {fq}")
            for ci, ticker in enumerate(TICKER_ORDER, 2):
                val = get_metric_value(df_src, ticker, fy, fq, metric_name)
                cell = ws.cell(row=data_row, column=ci)
                if val is None:
                    cell.value = "N/A"
                    cell.fill = MISSING_FILL
                    cell.alignment = Alignment(horizontal="center")
                elif unit_type == "USD":
                    cell.value = val / 1_000_000
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")
                    # check if derived
                    em_rows = df_src[(df_src["ticker"] == ticker) &
                                    (df_src["fiscal_year"] == fy) &
                                    (df_src["fiscal_quarter"] == fq) &
                                    (df_src["metric_name"] == metric_name)]
                    if not em_rows.empty and "extraction_method" in em_rows.columns:
                        em = em_rows.iloc[0].get("extraction_method", "")
                        if em in ("derived_ytd_difference",
                                  "fiscal_year_end_balance"):
                            cell.fill = DERIVED_FILL
                else:
                    cell.value = val
                    cell.number_format = '0.0%'
                    cell.alignment = Alignment(horizontal="right")
            data_row += 1

    # freeze on first header
    ws.freeze_panes = ws.cell(row=6, column=1)


def build_revenue_growth(wb, fin_df, met_df):
    _build_pivot_tab(
        wb, met_df, fin_df,
        "Revenue Growth",
        "Revenue & Growth Comparison",
        [
            ("Revenue", "USD", "fin"),
            ("YoY Revenue Growth", "%", "met"),
            ("TTM Revenue", "USD", "met"),
        ],
        None,
    )


def build_margins(wb, fin_df, met_df):
    _build_pivot_tab(
        wb, met_df, fin_df,
        "Margins",
        "Margin Comparison",
        [
            ("Gross Margin", "%", "met"),
            ("Operating Margin", "%", "met"),
            ("Free-Cash-Flow Margin", "%", "met"),
        ],
        None,
    )


def build_rd_intensity(wb, fin_df, met_df):
    _build_pivot_tab(
        wb, met_df, fin_df,
        "R&D Intensity",
        "R&D Spending Comparison",
        [
            ("R&D Expense", "USD", "fin"),
            ("R&D as % of Revenue", "%", "met"),
        ],
        None,
    )


def build_balance_sheet(wb, fin_df, met_df):
    _build_pivot_tab(
        wb, met_df, fin_df,
        "Balance Sheet",
        "Balance Sheet Comparison",
        [
            ("Cash and Cash Equivalents", "USD", "fin"),
            ("Total Debt", "USD", "fin"),
            ("Net Cash (Debt)", "USD", "met"),
        ],
        None,
    )


def build_validation(wb, fin_df, met_df, missing_df):
    """Tab 7: Validation — missing values, review flags, limitations."""
    ws = wb.create_sheet("Validation")
    write_title(ws, 1, "Validation & Data Quality")

    # ── Section 1: Missing values ──
    row = 3
    ws.cell(row=row, column=1,
            value="Missing Values (requires_manual_review)").font = SECTION_FONT
    row += 1
    headers = ["Ticker", "Company", "Metric", "Fiscal Year", "Quarter", "Reason"]
    widths = [8, 30, 22, 12, 8, 60]
    write_header_row(ws, row, headers, widths)
    row += 1

    # from financials: extraction_method == missing_requires_review
    missing_fin = fin_df[
        fin_df["extraction_method"] == "missing_requires_review"].copy()
    missing_fin = missing_fin.sort_values(["ticker", "metric_name",
                                           "fiscal_year", "fiscal_quarter"])
    for _, r in missing_fin.iterrows():
        ws.cell(row=row, column=1, value=r["ticker"])
        ws.cell(row=row, column=2, value=r["company"])
        ws.cell(row=row, column=3, value=r["metric_name"])
        ws.cell(row=row, column=4, value=r["fiscal_year"])
        ws.cell(row=row, column=5, value=r["fiscal_quarter"])
        reason = r.get("derived_from", "")
        ws.cell(row=row, column=6,
                value=reason if pd.notna(reason) else "")
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = MISSING_FILL
        row += 1

    # from metrics: empty values
    missing_met = met_df[met_df["value"].isna() | (met_df["value"] == "")].copy()
    missing_met = missing_met.sort_values(["ticker", "metric_name",
                                           "fiscal_year", "fiscal_quarter"])
    for _, r in missing_met.iterrows():
        ws.cell(row=row, column=1, value=r["ticker"])
        ws.cell(row=row, column=2, value=r["company"])
        ws.cell(row=row, column=3, value=r["metric_name"])
        ws.cell(row=row, column=4, value=r["fiscal_year"])
        ws.cell(row=row, column=5, value=r["fiscal_quarter"])
        formula = r.get("formula", "")
        ws.cell(row=row, column=6,
                value=formula if pd.notna(formula) else "")
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = MISSING_FILL
        row += 1

    # ── Section 2: Manual review flags (non-missing) ──
    row += 2
    ws.cell(row=row, column=1,
            value="Manual Review Flags (value present but unconfirmed)").font = SECTION_FONT
    row += 1
    headers2 = ["Ticker", "Metric", "Fiscal Year", "Quarter",
                "Value ($M)", "Extraction Method", "Notes"]
    write_header_row(ws, row, headers2, [8, 22, 12, 8, 16, 24, 50])
    row += 1

    review_rows = fin_df[
        (fin_df["requires_manual_review"] == "Yes") &
        (fin_df["extraction_method"] != "missing_requires_review")
    ].copy()
    review_rows = review_rows.sort_values(["ticker", "metric_name",
                                           "fiscal_year", "fiscal_quarter"])
    for _, r in review_rows.iterrows():
        ws.cell(row=row, column=1, value=r["ticker"])
        ws.cell(row=row, column=2, value=r["metric_name"])
        ws.cell(row=row, column=3, value=r["fiscal_year"])
        ws.cell(row=row, column=4, value=r["fiscal_quarter"])
        val = r["value"]
        try:
            ws.cell(row=row, column=5, value=float(val) / 1_000_000)
            ws.cell(row=row, column=5).number_format = '#,##0'
        except (ValueError, TypeError):
            ws.cell(row=row, column=5, value=val)
        ws.cell(row=row, column=6, value=r.get("extraction_method", ""))
        derived = r.get("derived_from", "")
        ws.cell(row=row, column=7,
                value=derived if pd.notna(derived) else "")
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = WARN_FILL
        row += 1

    # ── Section 3: Derived values summary ──
    row += 2
    ws.cell(row=row, column=1,
            value="Derived Values Summary").font = SECTION_FONT
    row += 1

    derived_methods = ["derived_ytd_difference", "fiscal_year_end_balance",
                       "reported_earnings_release"]
    for method in derived_methods:
        count = len(fin_df[fin_df["extraction_method"] == method])
        ws.cell(row=row, column=1, value=method)
        ws.cell(row=row, column=2, value=f"{count} entries across all companies")
        row += 1

    # ── Section 4: Company-specific limitations ──
    row += 2
    ws.cell(row=row, column=1,
            value="Company-Specific Limitations").font = SECTION_FONT
    row += 1
    for ticker in TICKER_ORDER:
        ws.cell(row=row, column=1, value=ticker).font = Font(bold=True)
        note_cell = ws.cell(row=row, column=2, value=COMPANY_NOTES[ticker])
        note_cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1

    # ── Section 5: Cross-company comparability warnings ──
    row += 2
    ws.cell(row=row, column=1,
            value="Cross-Company Comparability Warnings").font = SECTION_FONT
    row += 1
    warnings = [
        "Fiscal year-ends differ: QCOM=Sep, AMD=Dec, NVDA=Jan, INTC=Dec, AVGO=Nov. 'FY2025 Q1' refers to different calendar months across companies.",
        "Intel (IDM) has structurally higher CapEx and lower/negative FCF margins than fabless peers. Direct CapEx and FCF comparison is misleading without business-model adjustment.",
        "Intel cash tag includes restricted cash (~$447M / 3% overstatement vs peers using unrestricted cash).",
        "Broadcom Total Debt uses gross principal (~$2B / 3% above net carrying value used by other companies).",
        "Broadcom FY2024 is a 53-week VMware transition year with ~$16B acquired revenue and depressed margins from intangible amortization. Not comparable to FY2023 or peers.",
        "Qualcomm licensing revenue (QTL segment) inflates consolidated margins relative to pure semiconductor peers.",
        "Nvidia CapEx FY2024 Q1-Q3 sourced from 8-K press releases (pending manual confirmation). FCF and TTM FCF not calculated for those quarters.",
        "Diluted EPS Q4 missing for all 5 companies (15 instances). Cannot be derived from XBRL data.",
    ]
    for w in warnings:
        ws.cell(row=row, column=1, value="⚠")
        note_cell = ws.cell(row=row, column=2, value=w)
        note_cell.alignment = Alignment(wrap_text=True)
        note_cell.font = NOTE_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        row += 1


def build_data_dictionary(wb):
    """Tab 8: Data Dictionary — metric definitions and extraction methods."""
    ws = wb.create_sheet("Data Dictionary")
    write_title(ws, 1, "Data Dictionary")

    # Metric definitions
    row = 3
    ws.cell(row=row, column=1, value="Metric Definitions").font = SECTION_FONT
    row += 1
    headers = ["Metric", "Unit", "Definition", "Formula / Source"]
    widths = [28, 8, 50, 50]
    write_header_row(ws, row, headers, widths)
    row += 1

    metrics_def = [
        ("Revenue", "USD", "Total net revenue from SEC 10-Q/10-K filings", "XBRL tag varies by company (see tag map files)"),
        ("Cost of Revenue", "USD", "Total cost of goods sold / cost of revenue", "XBRL: CostOfRevenue or CostOfGoodsAndServicesSold"),
        ("Gross Profit", "USD", "Revenue minus Cost of Revenue", "Revenue - Cost of Revenue"),
        ("Operating Income", "USD", "Income from operations (GAAP)", "XBRL: OperatingIncomeLoss"),
        ("Net Income", "USD", "Net income attributable to parent (GAAP)", "XBRL: NetIncomeLoss or ProfitLoss (Broadcom)"),
        ("Diluted EPS", "USD/share", "Diluted earnings per share", "XBRL: EarningsPerShareDiluted. Q4 cannot be derived by subtraction."),
        ("R&D Expense", "USD", "Research and development expense", "XBRL: ResearchAndDevelopmentExpense"),
        ("Operating Cash Flow", "USD", "Net cash from operating activities", "XBRL: NetCashProvidedByUsedInOperatingActivities"),
        ("Capital Expenditure", "USD", "Purchases of property/equipment/productive assets", "XBRL tag varies: PaymentsToAcquireProductiveAssets or PaymentsToAcquirePropertyPlantAndEquipment"),
        ("Cash and Cash Equivalents", "USD", "Cash and equivalents at period end", "XBRL: CashAndCashEquivalentsAtCarryingValue (Intel includes restricted cash)"),
        ("Total Debt", "USD", "Total debt outstanding", "XBRL varies: single tag or sum of LongTermDebt + DebtCurrent. Broadcom uses gross principal."),
        ("Gross Margin", "%", "Gross Profit / Revenue", "Gross Profit / Revenue"),
        ("Operating Margin", "%", "Operating Income / Revenue", "Operating Income / Revenue"),
        ("Free Cash Flow", "USD", "Operating Cash Flow minus Capital Expenditure", "OCF - CapEx"),
        ("Free-Cash-Flow Margin", "%", "Free Cash Flow / Revenue", "FCF / Revenue"),
        ("R&D as % of Revenue", "%", "R&D Expense / Revenue", "R&D / Revenue"),
        ("Net Cash (Debt)", "USD", "Cash minus Total Debt", "Cash - Total Debt. Negative = net debt."),
        ("YoY Revenue Growth", "%", "Revenue change vs same quarter prior year", "(Current Q Revenue - Prior Year Q Revenue) / Prior Year Q Revenue"),
        ("TTM Revenue", "USD", "Trailing twelve months revenue", "Sum of last 4 quarters of Revenue"),
        ("TTM Operating Income", "USD", "Trailing twelve months operating income", "Sum of last 4 quarters of Operating Income"),
        ("TTM Free Cash Flow", "USD", "Trailing twelve months free cash flow", "Sum of last 4 quarters of FCF"),
    ]

    for name, unit, defn, formula in metrics_def:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=unit)
        ws.cell(row=row, column=3, value=defn).alignment = Alignment(
            wrap_text=True)
        ws.cell(row=row, column=4, value=formula).alignment = Alignment(
            wrap_text=True)
        row += 1

    # Extraction method labels
    row += 2
    ws.cell(row=row, column=1,
            value="Extraction Method Labels").font = SECTION_FONT
    row += 1
    headers2 = ["Label", "Meaning", "Cell Shading"]
    write_header_row(ws, row, headers2, [28, 60, 18])
    row += 1

    methods = [
        ("reported_standalone", "Value taken directly from a single quarterly SEC filing (10-Q or 10-K).", "None (white)"),
        ("derived_ytd_difference", "Q4 or standalone Q2/Q3 derived by subtracting YTD values. Q4 = FY (10-K) minus YTD Q3 (10-Q). Cash-flow Q2 = YTD Q2 minus Q1; Q3 = YTD Q3 minus YTD Q2.", "Yellow"),
        ("fiscal_year_end_balance", "Balance-sheet value at fiscal year-end, taken from the 10-K as the Q4 ending balance.", "Yellow"),
        ("reported_earnings_release", "Value sourced from an 8-K earnings press release rather than a 10-Q/10-K.", "Yellow (flagged for manual review)"),
        ("missing_requires_review", "Value could not be extracted or derived from available filings. Flagged for manual sourcing.", "Pink"),
        ("calculated", "Derived from other extracted metrics (margins, FCF, TTM, growth rates, net cash/debt).", "None (white)"),
    ]

    for label, meaning, shading in methods:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=meaning).alignment = Alignment(
            wrap_text=True)
        ws.cell(row=row, column=3, value=shading)
        row += 1


def build_source_audit(wb, fin_df):
    """Tab 9: Source Audit Trail — every financials row with source info."""
    ws = wb.create_sheet("Source Audit Trail")
    write_title(ws, 1, "Source Audit Trail")
    write_note(ws, 2, 1,
               "Every extracted financial metric with its SEC filing source, "
               "extraction method, and review status.")

    headers = ["Ticker", "Fiscal Year", "Quarter", "Metric",
               "Value ($M)", "Filing Date", "Form", "Source Reference",
               "Extraction Method", "Derived From", "Manual Review"]
    widths = [8, 12, 8, 22, 16, 12, 10, 60, 24, 50, 14]
    write_header_row(ws, 4, headers, widths)

    # sort
    q_order = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
    sorted_df = fin_df.copy()
    sorted_df["_ticker_rank"] = sorted_df["ticker"].map(
        {t: i for i, t in enumerate(TICKER_ORDER)})
    sorted_df["_q"] = sorted_df["fiscal_quarter"].map(q_order)
    sorted_df.sort_values(["_ticker_rank", "fiscal_year", "_q", "metric_name"],
                          ascending=[True, False, True, True], inplace=True)

    row = 5
    for _, r in sorted_df.iterrows():
        ws.cell(row=row, column=1, value=r["ticker"])
        ws.cell(row=row, column=2, value=r["fiscal_year"])
        ws.cell(row=row, column=3, value=r["fiscal_quarter"])
        ws.cell(row=row, column=4, value=r["metric_name"])

        val = r["value"]
        cell = ws.cell(row=row, column=5)
        if pd.isna(val) or val == "":
            cell.value = "N/A"
            cell.fill = MISSING_FILL
        else:
            try:
                cell.value = float(val) / 1_000_000
                cell.number_format = '#,##0'
            except (ValueError, TypeError):
                cell.value = val

        fd = r.get("filing_date", "")
        ws.cell(row=row, column=6, value=fd if pd.notna(fd) else "")
        ft = r.get("form_type", "")
        ws.cell(row=row, column=7, value=ft if pd.notna(ft) else "")
        sr = r.get("source_reference", "")
        ws.cell(row=row, column=8, value=sr if pd.notna(sr) else "")

        em = r.get("extraction_method", "")
        em_val = em if pd.notna(em) else ""
        ws.cell(row=row, column=9, value=em_val)

        df_val = r.get("derived_from", "")
        ws.cell(row=row, column=10, value=df_val if pd.notna(df_val) else "")

        mr = r.get("requires_manual_review", "")
        ws.cell(row=row, column=11, value=mr if pd.notna(mr) else "")

        # color coding
        if em_val == "missing_requires_review":
            for c in range(1, 12):
                ws.cell(row=row, column=c).fill = MISSING_FILL
        elif em_val in ("derived_ytd_difference", "fiscal_year_end_balance",
                        "reported_earnings_release"):
            for c in range(1, 12):
                ws.cell(row=row, column=c).fill = DERIVED_FILL
        if mr == "Yes" and em_val != "missing_requires_review":
            for c in range(1, 12):
                ws.cell(row=row, column=c).fill = WARN_FILL

        row += 1


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    print("Stage 10: Building five-company peer comparison model\n")

    # load data
    print("Loading financials CSVs...")
    fin_df = load_all_financials()
    print(f"  {len(fin_df)} total financials rows "
          f"({fin_df['ticker'].nunique()} companies)")

    print("Loading metrics CSVs...")
    met_df = load_all_metrics()
    print(f"  {len(met_df)} total metrics rows")

    print("Loading missing-metrics files...")
    missing_df = load_missing_metrics()
    print(f"  {len(missing_df)} missing-metric entries\n")

    # build consolidated CSV
    print("Building consolidated CSV...")
    consolidated = build_consolidated_csv(fin_df, met_df)

    # build Excel workbook
    print("\nBuilding Excel workbook...")
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    print("  Tab 1: Peer Summary")
    build_peer_summary(wb, fin_df, met_df)

    print("  Tab 2: Quarterly Trends")
    build_quarterly_trends(wb, fin_df, met_df)

    print("  Tab 3: Revenue Growth")
    build_revenue_growth(wb, fin_df, met_df)

    print("  Tab 4: Margins")
    build_margins(wb, fin_df, met_df)

    print("  Tab 5: R&D Intensity")
    build_rd_intensity(wb, fin_df, met_df)

    print("  Tab 6: Balance Sheet")
    build_balance_sheet(wb, fin_df, met_df)

    print("  Tab 7: Validation")
    build_validation(wb, fin_df, met_df, missing_df)

    print("  Tab 8: Data Dictionary")
    build_data_dictionary(wb)

    print("  Tab 9: Source Audit Trail")
    build_source_audit(wb, fin_df)

    wb.save(XLSX_OUT)
    print(f"\nWrote {XLSX_OUT}")

    # summary stats
    print("\n── Summary ──")
    for ticker in TICKER_ORDER:
        t_fin = fin_df[fin_df["ticker"] == ticker]
        t_met = met_df[met_df["ticker"] == ticker]
        missing_fin = t_fin[
            t_fin["extraction_method"] == "missing_requires_review"]
        missing_met = t_met[t_met["value"].isna() | (t_met["value"] == "")]
        review = t_fin[
            (t_fin["requires_manual_review"] == "Yes") &
            (t_fin["extraction_method"] != "missing_requires_review")]
        fy, fq = get_latest_quarter(fin_df, ticker)
        print(f"  {ticker}: {len(t_fin)} financials + {len(t_met)} metrics | "
              f"latest={f'FY{fy} {fq}'} | "
              f"missing={len(missing_fin) + len(missing_met)} | "
              f"manual_review={len(review)}")


if __name__ == "__main__":
    main()
