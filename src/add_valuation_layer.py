"""
Stage 10 — Valuation Layer

Reads:
  - data/manual_inputs/valuation_inputs.csv (market data + calculated EV)
  - data/processed/semiconductor_peer_metrics.csv (validated operating metrics)
  - data/processed/*_metrics.csv (TTM values per company)

Updates:
  - data/processed/semiconductor_peer_metrics.csv (appends valuation rows)
  - output/semiconductor_peer_comparison.xlsx (adds Valuation Snapshot tab,
    appends valuation summary to Peer Summary)

Does NOT modify company-level financials or metrics files.
Does NOT rerun SEC extraction pipelines.
"""

import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── paths ────────────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROCESSED_DIR = os.path.join(BASE_DIR, "data", "processed")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
INPUTS_DIR = os.path.join(BASE_DIR, "data", "manual_inputs")

PEER_CSV = os.path.join(PROCESSED_DIR, "semiconductor_peer_metrics.csv")
XLSX_PATH = os.path.join(OUTPUT_DIR, "semiconductor_peer_comparison.xlsx")
VAL_CSV = os.path.join(INPUTS_DIR, "valuation_inputs.csv")

# ── company metadata ─────────────────────────────────────────────────────────

TICKER_ORDER = ["QCOM", "AMD", "NVDA", "INTC", "AVGO"]

COMPANY_NAMES = {
    "QCOM": "Qualcomm Incorporated",
    "AMD": "Advanced Micro Devices Inc.",
    "NVDA": "Nvidia Corp.",
    "INTC": "Intel Corp.",
    "AVGO": "Broadcom Inc.",
}

FYEND_MAP = {
    "QCOM": "Sep (~Sep 29)",
    "AMD": "Dec (~Dec 28)",
    "NVDA": "Jan (~Jan 26)",
    "INTC": "Dec (~Dec 28)",
    "AVGO": "Nov (~Nov 3)",
}

MODEL_MAP = {
    "QCOM": "Fabless + Licensing",
    "AMD": "Fabless",
    "NVDA": "Fabless",
    "INTC": "IDM (Integrated Device Manufacturer)",
    "AVGO": "Fabless + Infrastructure Software",
}

# ── styles ───────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
TITLE_FONT = Font(bold=True, size=13)
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
NOTE_FONT = Font(italic=True, size=10, color="666666")
MANUAL_FILL = PatternFill("solid", fgColor="E2EFDA")
MISSING_FILL = PatternFill("solid", fgColor="F2DCDB")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))


def write_header_row(ws, row, headers, widths=None):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    if widths:
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w


def write_note(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = NOTE_FONT
    cell.alignment = Alignment(wrap_text=True)


# ── load data ────────────────────────────────────────────────────────────────

def load_valuation_inputs():
    df = pd.read_csv(VAL_CSV)
    return df


def get_ttm_value(met_df, ticker, metric_name):
    """Get the latest TTM value for a ticker."""
    rows = met_df[(met_df["ticker"] == ticker) &
                  (met_df["metric_name"] == metric_name)]
    rows = rows.dropna(subset=["value"])
    rows = rows[rows["value"] != ""]
    if rows.empty:
        return None
    q_map = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
    rows = rows.copy()
    rows["_q"] = rows["fiscal_quarter"].map(q_map)
    rows = rows.sort_values(["fiscal_year", "_q"], ascending=[False, False])
    try:
        return float(rows.iloc[0]["value"])
    except (ValueError, TypeError):
        return None


def get_latest_metric(met_df, ticker, metric_name):
    """Get the latest quarterly value for a ticker."""
    rows = met_df[(met_df["ticker"] == ticker) &
                  (met_df["metric_name"] == metric_name)]
    rows = rows.dropna(subset=["value"])
    rows = rows[rows["value"] != ""]
    if rows.empty:
        return None, None, None
    q_map = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
    rows = rows.copy()
    rows["_q"] = rows["fiscal_quarter"].map(q_map)
    rows = rows.sort_values(["fiscal_year", "_q"], ascending=[False, False])
    r = rows.iloc[0]
    try:
        return float(r["value"]), int(r["fiscal_year"]), r["fiscal_quarter"]
    except (ValueError, TypeError):
        return None, None, None


# ── compute valuation metrics ────────────────────────────────────────────────

def compute_valuation(val_df, met_df):
    """Compute valuation multiples for each company."""
    results = []
    for _, vr in val_df.iterrows():
        ticker = vr["ticker"]
        price = vr["share_price"]
        mcap = vr["market_cap"]
        ev = vr["enterprise_value"]
        cash = vr["cash"]
        debt = vr["total_debt"]
        debt_measure = vr["debt_measure"]

        ttm_rev = get_ttm_value(met_df, ticker, "TTM Revenue")
        ttm_fcf = get_ttm_value(met_df, ticker, "TTM Free Cash Flow")
        yoy_growth, yoy_fy, yoy_fq = get_latest_metric(
            met_df, ticker, "YoY Revenue Growth")
        net_cash_val, nc_fy, nc_fq = get_latest_metric(
            met_df, ticker, "Net Cash (Debt)")

        ev_ttm_rev = None
        if ev and ttm_rev and ttm_rev != 0:
            ev_ttm_rev = ev / ttm_rev

        p_ttm_fcf = None
        if mcap and ttm_fcf and ttm_fcf > 0:
            p_ttm_fcf = mcap / ttm_fcf

        fcf_yield = None
        if mcap and ttm_fcf and mcap > 0:
            fcf_yield = ttm_fcf / mcap

        results.append({
            "ticker": ticker,
            "share_price": price,
            "shares_outstanding": vr["shares_outstanding"],
            "shares_source_date": vr["shares_outstanding_source_date"],
            "market_cap": mcap,
            "enterprise_value": ev,
            "cash": cash,
            "total_debt": debt,
            "net_cash_debt": net_cash_val,
            "debt_measure": debt_measure,
            "ttm_revenue": ttm_rev,
            "ev_ttm_revenue": ev_ttm_rev,
            "ttm_fcf": ttm_fcf,
            "price_ttm_fcf": p_ttm_fcf,
            "fcf_yield": fcf_yield,
            "yoy_revenue_growth": yoy_growth,
            "yoy_growth_period": f"FY{yoy_fy} {yoy_fq}" if yoy_fy else None,
        })

    return results


# ── append valuation rows to peer CSV ────────────────────────────────────────

def append_valuation_to_csv(val_results, val_df):
    """Append valuation metrics to the consolidated peer CSV."""
    peer = pd.read_csv(PEER_CSV)

    # remove any prior valuation rows to allow reruns
    peer = peer[peer["extraction_method"] != "manual_valuation_input"]

    new_rows = []
    price_date = val_df.iloc[0]["share_price_date"]

    for vr in val_results:
        ticker = vr["ticker"]
        company = COMPANY_NAMES[ticker]
        base = {
            "company": company,
            "ticker": ticker,
            "fiscal_year": "",
            "fiscal_quarter": "",
            "extraction_method": "manual_valuation_input",
            "source_reference": f"Valuation snapshot {price_date}",
            "derived_from": "",
            "requires_manual_review": "",
            "formula": "",
        }

        metrics = [
            ("Share Price", vr["share_price"], "USD",
             f"Yahoo Finance closing price {price_date}"),
            ("Market Capitalization", vr["market_cap"], "USD",
             f"share_price ({vr['share_price']}) * shares_outstanding ({vr['shares_outstanding']:,.0f})"),
            ("Enterprise Value", vr["enterprise_value"], "USD",
             f"market_cap + total_debt - cash"),
            ("EV / TTM Revenue", vr["ev_ttm_revenue"], "x",
             f"EV / TTM Revenue" if vr["ev_ttm_revenue"] else ""),
            ("Price / TTM FCF", vr["price_ttm_fcf"], "x",
             f"Market Cap / TTM FCF" if vr["price_ttm_fcf"] else ""),
            ("FCF Yield", vr["fcf_yield"], "%",
             f"TTM FCF / Market Cap" if vr["fcf_yield"] else ""),
        ]

        for name, value, unit, formula in metrics:
            row = base.copy()
            row["metric_name"] = name
            row["value"] = value if value is not None else ""
            row["unit"] = unit
            row["formula"] = formula
            new_rows.append(row)

    new_df = pd.DataFrame(new_rows)
    combined = pd.concat([peer, new_df], ignore_index=True)
    combined.to_csv(PEER_CSV, index=False)
    print(f"Appended {len(new_rows)} valuation rows to {PEER_CSV}")
    print(f"Total rows now: {len(combined)}")


# ── Excel: Valuation Snapshot tab ────────────────────────────────────────────

def build_valuation_snapshot(wb, val_results, val_df):
    """Add Valuation Snapshot tab to the workbook."""
    ws = wb.create_sheet("Valuation Snapshot", 1)

    price_date = val_df.iloc[0]["share_price_date"]

    # title and notes
    ws.cell(row=1, column=1,
            value="Semiconductor Peer Valuation Snapshot").font = TITLE_FONT
    write_note(ws, 2, 1,
               f"Valuation as of: {price_date} (Yahoo Finance closing prices)")
    write_note(ws, 3, 1,
               "Market capitalization uses actual common shares outstanding "
               "from the latest available SEC filing cover page. "
               "Share-count source dates differ by company. "
               "Valuation snapshot prices use the June 4, 2026 closing price.")
    ws.merge_cells("A3:F3")

    # main valuation table
    row = 5
    headers = ["Metric"] + TICKER_ORDER
    widths = [28] + [22] * 5
    write_header_row(ws, row, headers, widths)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)

    # build lookup
    vr_map = {v["ticker"]: v for v in val_results}

    valuation_rows = [
        ("Share Price ($)", "share_price", "price", None),
        ("Shares Outstanding (M)", "shares_outstanding", "shares", None),
        ("Shares Source Date", "shares_source_date", "text", None),
        ("Market Cap ($M)", "market_cap", "usd_m", None),
        ("Enterprise Value ($M)", "enterprise_value", "usd_m", None),
        ("Cash ($M)", "cash", "usd_m", None),
        ("Total Debt ($M)", "total_debt", "usd_m", None),
        ("Net Cash / (Debt) ($M)", "net_cash_debt", "usd_m", None),
        ("Debt Measure", "debt_measure", "text", None),
    ]

    row = 6
    for label, key, fmt, _ in valuation_rows:
        ws.cell(row=row, column=1, value=label)
        for ci, ticker in enumerate(TICKER_ORDER, 2):
            vr = vr_map[ticker]
            val = vr.get(key)
            cell = ws.cell(row=row, column=ci)

            if val is None or (isinstance(val, float) and pd.isna(val)):
                cell.value = "N/A"
                cell.fill = MISSING_FILL
                cell.alignment = Alignment(horizontal="center")
            elif fmt == "price":
                cell.value = val
                cell.number_format = '#,##0.00'
                cell.fill = MANUAL_FILL
                cell.alignment = Alignment(horizontal="right")
            elif fmt == "shares":
                cell.value = val / 1_000_000
                cell.number_format = '#,##0.0'
                cell.fill = MANUAL_FILL
                cell.alignment = Alignment(horizontal="right")
            elif fmt == "usd_m":
                cell.value = val / 1_000_000
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif fmt == "text":
                cell.value = val
                cell.alignment = Alignment(horizontal="center")
                if key == "shares_source_date":
                    cell.fill = MANUAL_FILL
        row += 1

    # separator
    row += 1
    ws.cell(row=row, column=1,
            value="Valuation Multiples").font = SECTION_FONT
    row += 1

    multiple_rows = [
        ("TTM Revenue ($M)", "ttm_revenue", "usd_m"),
        ("EV / TTM Revenue", "ev_ttm_revenue", "multiple"),
        ("TTM Free Cash Flow ($M)", "ttm_fcf", "usd_m"),
        ("Price / TTM FCF", "price_ttm_fcf", "multiple"),
        ("FCF Yield", "fcf_yield", "pct"),
        ("YoY Revenue Growth", "yoy_revenue_growth", "pct"),
    ]

    for label, key, fmt in multiple_rows:
        ws.cell(row=row, column=1, value=label)
        for ci, ticker in enumerate(TICKER_ORDER, 2):
            vr = vr_map[ticker]
            val = vr.get(key)
            cell = ws.cell(row=row, column=ci)

            if val is None or (isinstance(val, float) and pd.isna(val)):
                cell.value = "N/A"
                cell.fill = MISSING_FILL
                cell.alignment = Alignment(horizontal="center")
            elif fmt == "usd_m":
                cell.value = val / 1_000_000
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif fmt == "multiple":
                cell.value = val
                cell.number_format = '0.0x'
                cell.alignment = Alignment(horizontal="right")
            elif fmt == "pct":
                cell.value = val
                cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal="right")
        row += 1

    # growth period label row
    ws.cell(row=row, column=1, value="Growth Period")
    for ci, ticker in enumerate(TICKER_ORDER, 2):
        vr = vr_map[ticker]
        gp = vr.get("yoy_growth_period")
        cell = ws.cell(row=row, column=ci, value=gp if gp else "N/A")
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(size=9, color="888888")
    row += 1

    # EV/Revenue vs Revenue Growth comparison
    row += 1
    ws.cell(row=row, column=1,
            value="EV / Revenue vs Growth").font = SECTION_FONT
    row += 1
    comp_headers = ["Ticker", "EV / TTM Rev", "YoY Rev Growth",
                    "EV/Rev per 1% Growth", "Business Model"]
    write_header_row(ws, row, comp_headers, [10, 16, 16, 22, 32])
    row += 1

    for ticker in TICKER_ORDER:
        vr = vr_map[ticker]
        ws.cell(row=row, column=1, value=ticker).font = Font(bold=True)

        ev_rev = vr.get("ev_ttm_revenue")
        growth = vr.get("yoy_revenue_growth")

        cell_ev = ws.cell(row=row, column=2)
        if ev_rev is not None:
            cell_ev.value = ev_rev
            cell_ev.number_format = '0.0x'
        else:
            cell_ev.value = "N/A"
            cell_ev.fill = MISSING_FILL

        cell_g = ws.cell(row=row, column=3)
        if growth is not None:
            cell_g.value = growth
            cell_g.number_format = '0.0%'
        else:
            cell_g.value = "N/A"
            cell_g.fill = MISSING_FILL

        cell_ratio = ws.cell(row=row, column=4)
        if ev_rev is not None and growth is not None and growth != 0:
            growth_pct = growth * 100
            if growth_pct > 0:
                cell_ratio.value = ev_rev / growth_pct
                cell_ratio.number_format = '0.00x'
            else:
                cell_ratio.value = "N/M (negative growth)"
                cell_ratio.fill = WARN_FILL
        else:
            cell_ratio.value = "N/A"
            cell_ratio.fill = MISSING_FILL

        ws.cell(row=row, column=5, value=MODEL_MAP[ticker])
        row += 1

    # notes section
    row += 2
    ws.cell(row=row, column=1,
            value="Notes & Limitations").font = SECTION_FONT
    row += 1

    notes = [
        "Green-shaded cells contain manually entered values (share price and shares outstanding from external sources).",
        "P/E ratio is not calculated (Q4 EPS missing for all companies; TTM EPS not available).",
        "EV / EBITDA is not calculated (EBITDA not yet derived; depreciation/amortization not extracted from SEC filings).",
        f"Intel cash includes restricted cash (~$447M / ~3% overstatement). debt_measure = carrying_value_includes_restricted_cash.",
        f"Broadcom Total Debt is gross principal (~$2B / ~3% above net carrying value). debt_measure = gross_principal.",
        "Broadcom share-count source date (2026-02-27) is older than peers (Apr-May 2026).",
        "Nvidia shares outstanding (24.2B) is a rounded figure from the SEC cover page.",
        "Intel Price/TTM FCF is not meaningful — TTM FCF is negative (IDM capital intensity).",
        "manually_reviewed = No. Values require user confirmation before use in external reports.",
    ]

    for note in notes:
        write_note(ws, row, 1, note)
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=6)
        row += 1


# ── Excel: append valuation summary to Peer Summary ─────────────────────────

def append_to_peer_summary(wb, val_results, val_df):
    """Add valuation metrics to the bottom of the Peer Summary tab."""
    ws = wb["Peer Summary"]
    max_row = ws.max_row

    price_date = val_df.iloc[0]["share_price_date"]
    vr_map = {v["ticker"]: v for v in val_results}

    row = max_row + 2
    ws.cell(row=row, column=1,
            value=f"Valuation Snapshot ({price_date})").font = SECTION_FONT
    row += 1
    write_note(ws, row, 1,
               "Green cells = manually entered market data. "
               "manually_reviewed = No.")
    row += 1

    summary_metrics = [
        ("Share Price ($)", "share_price", "price"),
        ("Market Cap ($M)", "market_cap", "usd_m"),
        ("Enterprise Value ($M)", "enterprise_value", "usd_m"),
        ("EV / TTM Revenue", "ev_ttm_revenue", "multiple"),
        ("Price / TTM FCF", "price_ttm_fcf", "multiple"),
        ("FCF Yield", "fcf_yield", "pct"),
    ]

    for label, key, fmt in summary_metrics:
        ws.cell(row=row, column=1, value=label)
        for ci, ticker in enumerate(TICKER_ORDER, 2):
            vr = vr_map[ticker]
            val = vr.get(key)
            cell = ws.cell(row=row, column=ci)

            if val is None or (isinstance(val, float) and pd.isna(val)):
                cell.value = "N/A"
                cell.fill = MISSING_FILL
                cell.alignment = Alignment(horizontal="center")
            elif fmt == "price":
                cell.value = val
                cell.number_format = '#,##0.00'
                cell.fill = MANUAL_FILL
                cell.alignment = Alignment(horizontal="right")
            elif fmt == "usd_m":
                cell.value = val / 1_000_000
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif fmt == "multiple":
                cell.value = val
                cell.number_format = '0.0x'
                cell.alignment = Alignment(horizontal="right")
            elif fmt == "pct":
                cell.value = val
                cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal="right")
        row += 1


# ── main ─────────────────────────────────────────────────────────────────────

def validate_inputs(val_df):
    """Validate that all required fields are populated."""
    errors = []
    dates = val_df["share_price_date"].unique()
    if len(dates) != 1:
        errors.append(f"share_price_date not uniform: {list(dates)}")

    for _, r in val_df.iterrows():
        t = r["ticker"]
        if pd.isna(r["share_price"]) or r["share_price"] == "":
            errors.append(f"{t}: missing share_price")
        if pd.isna(r["shares_outstanding"]) or r["shares_outstanding"] == "":
            errors.append(f"{t}: missing shares_outstanding")
        if pd.isna(r["source"]) or r["source"] == "":
            errors.append(f"{t}: missing source")
        if pd.isna(r["market_cap"]) or r["market_cap"] == "":
            errors.append(f"{t}: missing market_cap")
        if pd.isna(r["enterprise_value"]) or r["enterprise_value"] == "":
            errors.append(f"{t}: missing enterprise_value")

    return errors


def main():
    print("Stage 10 — Valuation Layer\n")

    # load inputs
    print("Loading valuation inputs...")
    val_df = load_valuation_inputs()
    print(f"  {len(val_df)} companies loaded")

    # validate
    print("\nValidating inputs...")
    errors = validate_inputs(val_df)
    if errors:
        print("  VALIDATION ERRORS:")
        for e in errors:
            print(f"    {e}")
        print("\nAborting. Fix valuation_inputs.csv and rerun.")
        return
    print("  All inputs valid")
    print(f"  Common price date: {val_df.iloc[0]['share_price_date']}")

    # load metrics for TTM values
    print("\nLoading metrics CSVs...")
    met_frames = []
    for path in glob.glob(os.path.join(PROCESSED_DIR, "*_metrics.csv")):
        df = pd.read_csv(path)
        ticker_map = {
            "Qualcomm Incorporated": "QCOM",
            "Advanced Micro Devices Inc.": "AMD",
            "Nvidia Corp.": "NVDA",
            "Intel Corp.": "INTC",
            "Broadcom Inc.": "AVGO",
        }
        df["ticker"] = df["company"].map(ticker_map)
        met_frames.append(df)
    met_df = pd.concat(met_frames, ignore_index=True)
    print(f"  {len(met_df)} metrics rows loaded")

    # compute valuation
    print("\nComputing valuation metrics...")
    val_results = compute_valuation(val_df, met_df)

    for vr in val_results:
        t = vr["ticker"]
        mcap_b = vr["market_cap"] / 1e9 if vr["market_cap"] else 0
        ev_b = vr["enterprise_value"] / 1e9 if vr["enterprise_value"] else 0
        ev_rev = vr["ev_ttm_revenue"]
        p_fcf = vr["price_ttm_fcf"]
        fcf_y = vr["fcf_yield"]
        print(f"  {t}: MCap=${mcap_b:.1f}B  EV=${ev_b:.1f}B  "
              f"EV/TTM_Rev={f'{ev_rev:.1f}x' if ev_rev else 'N/A'}  "
              f"P/TTM_FCF={f'{p_fcf:.1f}x' if p_fcf else 'N/A'}  "
              f"FCF_Yield={f'{fcf_y:.1%}' if fcf_y else 'N/A'}")

    # append to consolidated CSV
    print("\nUpdating consolidated CSV...")
    append_valuation_to_csv(val_results, val_df)

    # update Excel workbook
    print("\nUpdating Excel workbook...")
    wb = load_workbook(XLSX_PATH)

    # remove existing Valuation Snapshot if present (for reruns)
    if "Valuation Snapshot" in wb.sheetnames:
        del wb["Valuation Snapshot"]

    build_valuation_snapshot(wb, val_results, val_df)
    append_to_peer_summary(wb, val_results, val_df)

    wb.save(XLSX_PATH)
    print(f"Saved {XLSX_PATH}")

    # summary
    print("\n── Summary ──")
    print(f"  Price date: {val_df.iloc[0]['share_price_date']}")
    print(f"  Companies: {len(val_results)}")
    missing_items = []
    for vr in val_results:
        t = vr["ticker"]
        if vr["price_ttm_fcf"] is None:
            missing_items.append(f"{t}: Price/TTM FCF (TTM FCF unavailable "
                                 f"or negative)")
        if vr["fcf_yield"] is None:
            missing_items.append(f"{t}: FCF Yield (TTM FCF unavailable)")
    if missing_items:
        print(f"  Blank valuation fields:")
        for m in missing_items:
            print(f"    {m}")
    else:
        print(f"  All valuation fields populated")
    print(f"  manually_reviewed = No (pending user confirmation)")


if __name__ == "__main__":
    main()
