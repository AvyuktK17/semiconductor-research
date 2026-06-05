"""
export_excel.py
---------------
Reads a company's processed CSV and missing-metrics flag file,
then produces a clean multi-tab Excel workbook for analyst review.

Reads company metadata from config/companies.csv.

Tabs:
    1. Quarterly Financials — pivot grid: metrics as rows, quarters as columns
    2. Detail View          — full row-level data with all metadata columns
    3. Missing Metrics      — rows flagged for manual review + remaining gaps
    4. Data Dictionary      — column definitions and XBRL tag reference
    5. Manual Checks        — known assumptions and audit items

Output:  output/<short_id>_financial_history.xlsx

Usage:
    .venv/bin/python src/export_excel.py QCOM
    .venv/bin/python src/export_excel.py QCOM --test
"""

import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

from config_loader import load_company, load_xbrl_tags, parse_ticker_arg

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent
PROCESSED_DIR = PROJECT_ROOT / "data" / "processed"
MANUAL_CHECKS_DIR = PROJECT_ROOT / "data" / "manual_checks"
OUTPUT_DIR = PROJECT_ROOT / "output"

HEADER_FONT = Font(bold=True)
ITALIC_FONT = Font(italic=True)
DERIVED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                           fill_type="solid")  # light yellow
REVIEW_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC",
                          fill_type="solid")  # light pink

USD_FORMAT = '#,##0'
EPS_FORMAT = '#,##0.00'
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_BOTTOM = Alignment(wrap_text=True, vertical="bottom")


def is_test_mode() -> bool:
    return "--test" in sys.argv


def resolve_paths(short_id: str) -> tuple[Path, Path, Path]:
    """Return (processed_dir, manual_checks_dir, output_dir)."""
    if is_test_mode():
        data_base = PROJECT_ROOT / "data" / "test_outputs" / short_id
        out_base = PROJECT_ROOT / "output" / "test_outputs" / short_id
        return data_base, data_base, out_base
    return PROCESSED_DIR, MANUAL_CHECKS_DIR, OUTPUT_DIR


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def find_latest_file(directory: Path, pattern: str) -> Path:
    candidates = sorted(directory.glob(pattern))
    if not candidates:
        sys.exit(f"ERROR: No files matching '{pattern}' in {directory}")
    return candidates[-1]


def set_column_widths(ws, widths: dict):
    for col_num, width in widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width


def style_header_row(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Tab 1: Quarterly Financials (pivot grid)
# ---------------------------------------------------------------------------

METRIC_ORDER = [
    "Revenue",
    "Cost of Revenue",
    "Gross Profit",
    "Operating Income",
    "Net Income",
    "Diluted EPS",
    "R&D Expense",
    "Operating Cash Flow",
    "Capital Expenditure",
    "Cash and Cash Equivalents",
    "Total Debt",
]


def build_quarterly_financials(wb, rows: list[dict]):
    ws = wb.active
    ws.title = "Quarterly Financials"

    q_sort = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
    seen = {}
    for r in rows:
        key = (int(r["fiscal_year"]), r["fiscal_quarter"])
        if key not in seen:
            seen[key] = r

    periods = sorted(seen.keys(), key=lambda k: (k[0], q_sort.get(k[1], 9)))

    lookup = {}
    for r in rows:
        key = (r["metric_name"], int(r["fiscal_year"]), r["fiscal_quarter"])
        lookup[key] = r

    # --- Headers ---
    headers = ["Metric", "Unit"]
    for fy, fq in periods:
        headers.append(f"FY{fy} {fq}")
    ws.append(headers)

    # --- Data rows ---
    for metric in METRIC_ORDER:
        sample = next((lookup.get((metric, p[0], p[1]))
                       for p in periods
                       if (metric, p[0], p[1]) in lookup), None)
        unit = sample["unit"] if sample else "USD"

        row_data = [metric, unit]
        for fy, fq in periods:
            entry = lookup.get((metric, fy, fq))
            if entry and entry["value"]:
                row_data.append(float(entry["value"]))
            else:
                row_data.append("")
        ws.append(row_data)

    # --- Formatting ---
    style_header_row(ws)
    for cell in ws[1]:
        cell.alignment = WRAP_BOTTOM

    widths = {1: 28, 2: 12}
    for i in range(3, 3 + len(periods)):
        widths[i] = 16
    set_column_widths(ws, widths)

    for row_idx in range(2, ws.max_row + 1):
        metric_name = ws.cell(row=row_idx, column=1).value
        for col_idx in range(3, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            fy, fq = periods[col_idx - 3]
            entry = lookup.get((metric_name, fy, fq))

            if isinstance(cell.value, (int, float)):
                if metric_name == "Diluted EPS":
                    cell.number_format = EPS_FORMAT
                else:
                    cell.number_format = USD_FORMAT

            if entry:
                if entry["requires_manual_review"] == "Yes":
                    cell.fill = REVIEW_FILL
                elif entry["extraction_method"] in (
                        "derived_ytd_difference", "fiscal_year_end_balance"):
                    cell.fill = DERIVED_FILL

    # Legend row
    ws.append([])
    legend_row = ws.max_row + 1
    ws.cell(row=legend_row, column=1, value="Legend:").font = ITALIC_FONT
    ws.cell(row=legend_row, column=2,
            value="Yellow = derived value").fill = DERIVED_FILL
    ws.cell(row=legend_row, column=3,
            value="Pink = requires manual review").fill = REVIEW_FILL

    ws.cell(row=legend_row + 1, column=1,
            value="Source: SEC EDGAR XBRL company-facts API").font = ITALIC_FONT


# ---------------------------------------------------------------------------
# Tab 1b: Detail View (full row-level data)
# ---------------------------------------------------------------------------

def build_detail_view(wb, rows: list[dict]):
    ws = wb.create_sheet("Detail View")

    detail_headers = [
        "Fiscal Year", "Fiscal Quarter", "Metric", "Value", "Unit",
        "Extraction Method", "Derived From", "Requires Manual Review",
        "Filing Date", "Form Type", "Source Reference",
    ]
    ws.append(detail_headers)

    q_sort = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}
    sorted_rows = sorted(rows, key=lambda r: (
        r["metric_name"],
        int(r["fiscal_year"]),
        q_sort.get(r["fiscal_quarter"], 9),
    ))

    for r in sorted_rows:
        val = float(r["value"]) if r["value"] else ""
        ws.append([
            int(r["fiscal_year"]),
            r["fiscal_quarter"],
            r["metric_name"],
            val,
            r["unit"],
            r["extraction_method"],
            r["derived_from"],
            r["requires_manual_review"],
            r["filing_date"],
            r["form_type"],
            r["source_reference"],
        ])

    style_header_row(ws)
    set_column_widths(ws, {
        1: 12, 2: 14, 3: 26, 4: 18, 5: 12,
        6: 24, 7: 48, 8: 22, 9: 12, 10: 14, 11: 60,
    })

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=4)
        metric = ws.cell(row=row_idx, column=3).value
        if isinstance(cell.value, (int, float)):
            cell.number_format = (EPS_FORMAT if metric == "Diluted EPS"
                                  else USD_FORMAT)

        if ws.cell(row=row_idx, column=8).value == "Yes":
            for col in range(1, 12):
                ws.cell(row=row_idx, column=col).fill = REVIEW_FILL
        elif ws.cell(row=row_idx, column=6).value in (
                "derived_ytd_difference", "fiscal_year_end_balance"):
            for col in range(1, 12):
                ws.cell(row=row_idx, column=col).fill = DERIVED_FILL


# ---------------------------------------------------------------------------
# Tab 2: Missing Metrics
# ---------------------------------------------------------------------------

def build_missing_metrics(wb, missing_path: Path, rows: list[dict]):
    ws = wb.create_sheet("Missing Metrics")

    ws.append(["Metric", "Fiscal Year", "Fiscal Quarter",
               "Extraction Method", "Reason"])

    with open(missing_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            ws.append([
                r["metric"],
                r.get("fiscal_year", ""),
                r.get("fiscal_quarter", ""),
                r.get("extraction_method", ""),
                r.get("reason", ""),
            ])

    remaining_gaps = [
        ["Free Cash Flow", "", "", "not_yet_implemented",
         "Derived metric: Operating Cash Flow minus Capital Expenditure"],
        ["Gross Margin", "", "", "not_yet_implemented",
         "Ratio: Gross Profit / Revenue"],
        ["Operating Margin", "", "", "not_yet_implemented",
         "Ratio: Operating Income / Revenue"],
        ["Free-Cash-Flow Margin", "", "", "not_yet_implemented",
         "Ratio: Free Cash Flow / Revenue"],
        ["R&D as % of Revenue", "", "", "not_yet_implemented",
         "Ratio: R&D Expense / Revenue"],
    ]

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        existing.add(row[0])

    for gap in remaining_gaps:
        if gap[0] not in existing:
            ws.append(gap)

    style_header_row(ws)
    set_column_widths(ws, {1: 24, 2: 12, 3: 16, 4: 24, 5: 58})


# ---------------------------------------------------------------------------
# Tab 3: Data Dictionary
# ---------------------------------------------------------------------------

def build_data_dictionary(wb, company_name: str, tags_config: dict):
    ws = wb.create_sheet("Data Dictionary")

    ws.append(["Field", "Description"])

    dictionary = [
        ["company", f"Legal entity name ({company_name})"],
        ["fiscal_year", "Company fiscal year (see company config for FY-end date)"],
        ["fiscal_quarter",
         "Q1-Q4; Q4 is derived if the company does not file a standalone Q4 10-Q"],
        ["metric_name", "Financial metric label (see Metric Reference below)"],
        ["value", "Reported or derived value in the unit shown"],
        ["unit", "USD or USD/shares"],
        ["filing_date", "Date the SEC filing was submitted"],
        ["form_type", "10-Q, 10-K, or '10-K + 10-Q' for derived Q4 values"],
        ["source_reference", "URL to the specific SEC filing on EDGAR"],
        ["extraction_method",
         "How the value was obtained (see Extraction Methods below)"],
        ["derived_from",
         "Formula and source values used for derived metrics"],
        ["requires_manual_review",
         "'Yes' if the value could not be extracted or verified"],
    ]

    for row in dictionary:
        ws.append(row)

    # Extraction methods
    ws.append([])
    ws.append(["Extraction Method", "Meaning"])
    methods = [
        ["reported_standalone",
         "Value taken directly from a single-quarter SEC filing entry"],
        ["derived_ytd_difference",
         "Calculated by subtracting cumulative YTD periods "
         "(e.g., Q4 = FY minus 9M YTD; Q2 = 6M YTD minus Q1)"],
        ["fiscal_year_end_balance",
         "Balance-sheet snapshot from the 10-K fiscal year-end date, "
         "used as the Q4 ending balance"],
        ["missing_requires_review",
         "Value could not be extracted or safely derived; "
         "flagged for manual review"],
    ]
    for row in methods:
        ws.append(row)

    # Metric reference — built from the XBRL tag config
    ws.append([])
    ws.append(["Metric Reference", "XBRL Tag(s) Used"])

    for metric_name, cfg in tags_config["metric_map"].items():
        tag = cfg["tag"]
        if isinstance(tag, list):
            tag_str = " + ".join(f"us-gaap:{t}" for t in tag)
        else:
            tag_str = f"us-gaap:{tag}"
        ws.append([metric_name, tag_str])

    for metric_name, cfg in tags_config.get("derived_metrics", {}).items():
        ws.append([metric_name, f"Derived: {cfg['formula']}"])

    style_header_row(ws)
    set_column_widths(ws, {1: 30, 2: 60})

    for row in ws.iter_rows():
        if row[0].value in ("Extraction Method", "Metric Reference"):
            row[0].font = HEADER_FONT
            row[1].font = HEADER_FONT


# ---------------------------------------------------------------------------
# Tab 4: Manual Checks
# ---------------------------------------------------------------------------

def build_manual_checks(wb, company_name: str, ticker: str,
                        rows: list[dict]):
    ws = wb.create_sheet("Manual Checks")

    ws.append(["Item", "Detail", "Action Required"])

    checks = [
        [
            "Q4 values are derived",
            f"{company_name} does not file a Q4 10-Q. All Q4 duration "
            "metrics are calculated as FY (10-K) minus 9-month YTD "
            "(Q3 10-Q). Q4 balance-sheet items use the fiscal-year-end "
            "snapshot.",
            f"Cross-check derived Q4 values against {company_name}'s "
            "earnings press releases for each fiscal year",
        ],
        [
            "Cash-flow Q2/Q3 are derived",
            "SEC XBRL only reports cumulative YTD for cash-flow items. "
            "Q2 = YTD_Q2 minus Q1; Q3 = YTD_Q3 minus YTD_Q2.",
            "Verify a sample quarter against the 10-Q filing directly",
        ],
        [
            "Gross Profit is derived",
            f"{company_name} does not file a GrossProfit XBRL tag. "
            "Calculated as Revenue minus CostOfRevenue each quarter.",
            "Verify against the income statement in a recent 10-Q",
        ],
        [
            "Total Debt is a sum",
            "Total Debt = LongTermDebt + DebtCurrent. Most companies "
            "do not file a single TotalDebt XBRL tag.",
            "Verify the sum matches the balance sheet in a recent 10-Q/10-K",
        ],
        [
            "Capital Expenditure tag",
            "The XBRL tag used for CapEx varies by company. Check "
            "the Data Dictionary tab for the specific tag used.",
            "Cross-check against the cash flow statement in a recent 10-Q",
        ],
        [
            "Diluted EPS Q4 not derived",
            "EPS cannot be derived by subtracting YTD values because "
            "diluted share counts change across quarters. Q4 EPS is "
            "left blank and flagged for manual review.",
            f"Source Q4 EPS from {company_name}'s earnings press release",
        ],
    ]

    # Qualcomm-specific note
    if ticker == "QCOM":
        checks.append([
            "FY2025 Q4 Net Income is negative",
            "The derived Q4 value is -$3.1B (FY total $5.5B minus "
            "9-month YTD $8.7B). This likely reflects a one-time "
            "charge or impairment.",
            "Verify against the FY2025 10-K notes and earnings release",
        ])

    for row in checks:
        ws.append(row)

    style_header_row(ws)
    set_column_widths(ws, {1: 32, 2: 58, 3: 52})

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_TOP


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ticker = parse_ticker_arg()
    company = load_company(ticker)

    short_id = company["short_identifier"]
    company_name = company["company_name"]
    tags_config = load_xbrl_tags(short_id)

    proc_dir, checks_dir, out_dir = resolve_paths(short_id)

    csv_path = find_latest_file(proc_dir,
                                f"{short_id}_financials_*.csv")
    missing_path = find_latest_file(checks_dir,
                                    f"{short_id}_missing_metrics_*.csv")

    print(f"Reading: {csv_path.name}")
    print(f"Reading: {missing_path.name}")

    with open(csv_path, "r", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    wb = openpyxl.Workbook()

    build_quarterly_financials(wb, rows)
    build_detail_view(wb, rows)
    build_missing_metrics(wb, missing_path, rows)
    build_data_dictionary(wb, company_name, tags_config)
    build_manual_checks(wb, company_name, ticker, rows)

    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{short_id}_financial_history.xlsx"
    wb.save(out_path)

    print(f"\nWorkbook saved to: {out_path}")
    print(f"Tabs: {wb.sheetnames}")

    ws1 = wb["Quarterly Financials"]
    print(f"Quarterly Financials: {ws1.max_row - 1} rows x "
          f"{ws1.max_column - 2} quarter columns")

    ws2 = wb["Detail View"]
    print(f"Detail View: {ws2.max_row - 1} data rows")


if __name__ == "__main__":
    main()
