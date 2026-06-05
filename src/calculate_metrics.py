"""
calculate_metrics.py
--------------------
Calculates derived financial metrics from the validated Qualcomm
quarterly dataset and appends a new tab to the Excel workbook.

Formulas:
    Gross Margin %              = Gross Profit / Revenue * 100
    Operating Margin %          = Operating Income / Revenue * 100
    Free Cash Flow              = Operating Cash Flow - Capital Expenditure
    Free-Cash-Flow Margin %     = Free Cash Flow / Revenue * 100
    R&D as % of Revenue         = R&D Expense / Revenue * 100
    Net Cash (Debt)             = Cash and Cash Equivalents - Total Debt
                                  (positive = net cash, negative = net debt)
    YoY Revenue Growth %        = (Revenue_Q - Revenue_Q-4) / |Revenue_Q-4| * 100
                                  where Q-4 is the same quarter one year prior
    TTM Revenue                 = sum of the most recent 4 quarters of Revenue
    TTM Operating Income        = sum of the most recent 4 quarters of Op Income
    TTM Free Cash Flow          = sum of the most recent 4 quarters of FCF

Reads:   data/processed/qualcomm_financials_<date>.csv
Writes:  data/processed/qualcomm_metrics.csv
Updates: output/qualcomm_financial_history.xlsx (adds "Calculated Metrics" tab)

Usage:
    .venv/bin/python src/calculate_metrics.py
"""

import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent
PROCESSED_DIR = PROJECT_ROOT / "data" / "processed"
OUTPUT_DIR = PROJECT_ROOT / "output"

HEADER_FONT = Font(bold=True)
WRAP_BOTTOM = Alignment(wrap_text=True, vertical="bottom")

PCT_FORMAT = '0.0%'
USD_FORMAT = '#,##0'

Q_ORDER = ["Q1", "Q2", "Q3", "Q4"]

CSV_COLUMNS = [
    "company",
    "fiscal_year",
    "fiscal_quarter",
    "metric_name",
    "value",
    "unit",
    "formula",
]


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def find_latest_file(directory: Path, pattern: str) -> Path:
    candidates = sorted(directory.glob(pattern))
    if not candidates:
        sys.exit(f"ERROR: No files matching '{pattern}' in {directory}")
    return candidates[-1]


def load_base_data(csv_path: Path) -> dict:
    """
    Load the processed CSV into a lookup dict:
        {(metric_name, fiscal_year, fiscal_quarter): float_value}
    Skips rows with empty values.
    """
    lookup = {}
    with open(csv_path, "r", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if r["value"]:
                key = (r["metric_name"], int(r["fiscal_year"]),
                       r["fiscal_quarter"])
                lookup[key] = float(r["value"])
    return lookup


def get_val(lookup, metric, fy, fq):
    """Return value or None if missing."""
    return lookup.get((metric, fy, fq))


def all_periods(lookup) -> list[tuple[int, str]]:
    """Return sorted list of (fiscal_year, fiscal_quarter) in the data."""
    periods = set()
    for (_, fy, fq) in lookup:
        periods.add((fy, fq))
    q_idx = {q: i for i, q in enumerate(Q_ORDER)}
    return sorted(periods, key=lambda p: (p[0], q_idx.get(p[1], 99)))


def prior_quarter(fy: int, fq: str, offset: int = 1):
    """Return (fy, fq) for `offset` quarters earlier."""
    idx = Q_ORDER.index(fq)
    for _ in range(offset):
        idx -= 1
        if idx < 0:
            idx = 3
            fy -= 1
    return fy, Q_ORDER[idx]


# ---------------------------------------------------------------------------
# Metric calculations
# ---------------------------------------------------------------------------

def calculate_all(lookup: dict) -> list[dict]:
    """Calculate all derived metrics and return as a list of row dicts."""
    periods = all_periods(lookup)
    rows = []

    def add(fy, fq, name, value, unit, formula):
        rows.append({
            "company": "Qualcomm Incorporated",
            "fiscal_year": fy,
            "fiscal_quarter": fq,
            "metric_name": name,
            "value": round(value, 2) if value is not None else "",
            "unit": unit,
            "formula": formula,
        })

    for fy, fq in periods:
        rev = get_val(lookup, "Revenue", fy, fq)
        gp = get_val(lookup, "Gross Profit", fy, fq)
        oi = get_val(lookup, "Operating Income", fy, fq)
        ocf = get_val(lookup, "Operating Cash Flow", fy, fq)
        capex = get_val(lookup, "Capital Expenditure", fy, fq)
        rnd = get_val(lookup, "R&D Expense", fy, fq)
        cash = get_val(lookup, "Cash and Cash Equivalents", fy, fq)
        debt = get_val(lookup, "Total Debt", fy, fq)

        # --- Gross Margin ---
        if gp is not None and rev is not None and rev != 0:
            gm = gp / rev
            add(fy, fq, "Gross Margin", gm, "%",
                f"Gross Profit ({gp:,.0f}) / Revenue ({rev:,.0f})")
        else:
            add(fy, fq, "Gross Margin", None, "%",
                "Missing: Gross Profit or Revenue")

        # --- Operating Margin ---
        if oi is not None and rev is not None and rev != 0:
            om = oi / rev
            add(fy, fq, "Operating Margin", om, "%",
                f"Operating Income ({oi:,.0f}) / Revenue ({rev:,.0f})")
        else:
            add(fy, fq, "Operating Margin", None, "%",
                "Missing: Operating Income or Revenue")

        # --- Free Cash Flow ---
        if ocf is not None and capex is not None:
            fcf = ocf - capex
            add(fy, fq, "Free Cash Flow", fcf, "USD",
                f"Operating Cash Flow ({ocf:,.0f}) - "
                f"Capital Expenditure ({capex:,.0f})")
        else:
            add(fy, fq, "Free Cash Flow", None, "USD",
                "Missing: OCF or CapEx")

        # --- FCF Margin ---
        if ocf is not None and capex is not None and rev is not None and rev != 0:
            fcf = ocf - capex
            fcf_margin = fcf / rev
            add(fy, fq, "Free-Cash-Flow Margin", fcf_margin, "%",
                f"FCF ({fcf:,.0f}) / Revenue ({rev:,.0f})")
        else:
            add(fy, fq, "Free-Cash-Flow Margin", None, "%",
                "Missing: FCF components or Revenue")

        # --- R&D as % of Revenue ---
        if rnd is not None and rev is not None and rev != 0:
            rnd_pct = rnd / rev
            add(fy, fq, "R&D as % of Revenue", rnd_pct, "%",
                f"R&D Expense ({rnd:,.0f}) / Revenue ({rev:,.0f})")
        else:
            add(fy, fq, "R&D as % of Revenue", None, "%",
                "Missing: R&D Expense or Revenue")

        # --- Net Cash (Debt) ---
        if cash is not None and debt is not None:
            net = cash - debt
            label = "net cash" if net >= 0 else "net debt"
            add(fy, fq, "Net Cash (Debt)", net, "USD",
                f"Cash ({cash:,.0f}) - Total Debt ({debt:,.0f}) = {label}")
        else:
            add(fy, fq, "Net Cash (Debt)", None, "USD",
                "Missing: Cash or Total Debt")

        # --- YoY Revenue Growth ---
        py_fy, py_fq = prior_quarter(fy, fq, offset=4)
        rev_prior = get_val(lookup, "Revenue", py_fy, py_fq)
        if rev is not None and rev_prior is not None and rev_prior != 0:
            yoy = (rev - rev_prior) / abs(rev_prior)
            add(fy, fq, "YoY Revenue Growth", yoy, "%",
                f"(Revenue {rev:,.0f} - Prior {rev_prior:,.0f}) "
                f"/ |Prior {rev_prior:,.0f}|")
        else:
            add(fy, fq, "YoY Revenue Growth", None, "%",
                f"Missing: current or prior-year Revenue "
                f"(FY{py_fy} {py_fq})")

        # --- TTM metrics (need 4 trailing quarters) ---
        ttm_quarters = [(fy, fq)]
        for i in range(1, 4):
            ttm_quarters.append(prior_quarter(fy, fq, offset=i))

        for ttm_name, base_metric in [
            ("TTM Revenue", "Revenue"),
            ("TTM Operating Income", "Operating Income"),
        ]:
            vals = [get_val(lookup, base_metric, y, q) for y, q in ttm_quarters]
            if all(v is not None for v in vals):
                ttm = sum(vals)
                quarters_str = " + ".join(
                    f"FY{y} {q}" for y, q in reversed(ttm_quarters))
                add(fy, fq, ttm_name, ttm, "USD",
                    f"Sum of {quarters_str}")
            else:
                add(fy, fq, ttm_name, None, "USD",
                    "Missing: not enough trailing quarters")

        # TTM FCF — needs OCF and CapEx for each of 4 quarters
        ocf_vals = [get_val(lookup, "Operating Cash Flow", y, q)
                    for y, q in ttm_quarters]
        capex_vals = [get_val(lookup, "Capital Expenditure", y, q)
                      for y, q in ttm_quarters]
        if (all(v is not None for v in ocf_vals) and
                all(v is not None for v in capex_vals)):
            ttm_fcf = sum(o - c for o, c in zip(ocf_vals, capex_vals))
            quarters_str = " + ".join(
                f"FY{y} {q}" for y, q in reversed(ttm_quarters))
            add(fy, fq, "TTM Free Cash Flow", ttm_fcf, "USD",
                f"Sum of quarterly FCF for {quarters_str}")
        else:
            add(fy, fq, "TTM Free Cash Flow", None, "USD",
                "Missing: not enough trailing OCF/CapEx quarters")

    return rows


# ---------------------------------------------------------------------------
# Excel tab
# ---------------------------------------------------------------------------

METRIC_DISPLAY_ORDER = [
    "Gross Margin",
    "Operating Margin",
    "Free Cash Flow",
    "Free-Cash-Flow Margin",
    "R&D as % of Revenue",
    "Net Cash (Debt)",
    "YoY Revenue Growth",
    "TTM Revenue",
    "TTM Operating Income",
    "TTM Free Cash Flow",
]


def add_calculated_metrics_tab(wb_path: Path, calc_rows: list[dict]):
    """Open existing workbook, add/replace a Calculated Metrics tab."""
    wb = openpyxl.load_workbook(wb_path)

    if "Calculated Metrics" in wb.sheetnames:
        del wb["Calculated Metrics"]

    ws = wb.create_sheet("Calculated Metrics", 1)

    # Build lookup and periods
    q_idx = {q: i for i, q in enumerate(Q_ORDER)}
    lookup = {}
    period_set = set()
    for r in calc_rows:
        if r["value"] != "":
            key = (r["metric_name"], r["fiscal_year"], r["fiscal_quarter"])
            lookup[key] = r
        period_set.add((r["fiscal_year"], r["fiscal_quarter"]))

    periods = sorted(period_set, key=lambda p: (p[0], q_idx.get(p[1], 99)))

    # Headers
    headers = ["Metric", "Unit"]
    for fy, fq in periods:
        headers.append(f"FY{fy} {fq}")
    ws.append(headers)

    # Data rows
    for metric in METRIC_DISPLAY_ORDER:
        sample = next((lookup.get((metric, p[0], p[1]))
                        for p in periods
                        if (metric, p[0], p[1]) in lookup), None)
        unit = sample["unit"] if sample else ""
        row_data = [metric, unit]

        for fy, fq in periods:
            entry = lookup.get((metric, fy, fq))
            if entry and entry["value"] != "":
                row_data.append(entry["value"])
            else:
                row_data.append("")
        ws.append(row_data)

    # Formatting
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = WRAP_BOTTOM
    ws.freeze_panes = "A2"

    widths = {1: 26, 2: 8}
    for i in range(3, 3 + len(periods)):
        widths[i] = 16
    for col_num, width in widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    for row_idx in range(2, ws.max_row + 1):
        unit_cell = ws.cell(row=row_idx, column=2).value
        for col_idx in range(3, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                if unit_cell == "%":
                    cell.number_format = PCT_FORMAT
                else:
                    cell.number_format = USD_FORMAT

    wb.save(wb_path)
    return wb.sheetnames


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    csv_path = find_latest_file(PROCESSED_DIR, "qualcomm_financials_*.csv")
    print(f"Reading: {csv_path.name}")

    lookup = load_base_data(csv_path)
    print(f"Loaded {len(lookup)} base data points")

    calc_rows = calculate_all(lookup)

    # Write metrics CSV
    out_csv = PROCESSED_DIR / "qualcomm_metrics.csv"
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
        writer.writeheader()
        writer.writerows(calc_rows)

    filled = sum(1 for r in calc_rows if r["value"] != "")
    empty = sum(1 for r in calc_rows if r["value"] == "")
    print(f"Saved {len(calc_rows)} rows ({filled} calculated, "
          f"{empty} missing) to: {out_csv}")

    # Update Excel workbook
    xlsx_path = OUTPUT_DIR / "qualcomm_financial_history.xlsx"
    if not xlsx_path.exists():
        sys.exit(f"ERROR: Workbook not found at {xlsx_path}. "
                 "Run export_excel.py first.")

    tabs = add_calculated_metrics_tab(xlsx_path, calc_rows)
    print(f"Updated workbook tabs: {tabs}")


if __name__ == "__main__":
    main()
