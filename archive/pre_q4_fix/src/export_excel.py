"""
export_excel.py
---------------
Reads the processed Qualcomm CSV and missing-metrics flag file,
then produces a clean multi-tab Excel workbook for analyst review.

Tabs:
    1. Quarterly Financials — pivot of metrics by quarter
    2. Missing Metrics      — metrics we could not extract
    3. Data Dictionary      — explains every column and metric
    4. Manual Checks        — assumptions and items needing review

Output:  output/qualcomm_financial_history.xlsx

Usage:
    .venv/bin/python src/export_excel.py
"""

import csv
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, numbers

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent
PROCESSED_DIR = PROJECT_ROOT / "data" / "processed"
MANUAL_CHECKS_DIR = PROJECT_ROOT / "data" / "manual_checks"
OUTPUT_DIR = PROJECT_ROOT / "output"

HEADER_FONT = Font(bold=True)

# USD columns get comma-separated thousands with no decimals.
USD_FORMAT = '#,##0'
# EPS gets two decimal places.
EPS_FORMAT = '#,##0.00'


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def find_latest_file(directory: Path, pattern: str) -> Path:
    candidates = sorted(directory.glob(pattern))
    if not candidates:
        sys.exit(f"ERROR: No files matching '{pattern}' in {directory}")
    return candidates[-1]


def set_column_widths(ws, widths: dict):
    """Set column widths by column number (1-indexed)."""
    for col_num, width in widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width


def style_header_row(ws):
    """Bold the first row and freeze it."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Tab 1: Quarterly Financials (pivoted)
# ---------------------------------------------------------------------------

def build_quarterly_financials(wb, rows: list[dict]):
    ws = wb.active
    ws.title = "Quarterly Financials"

    # Use Revenue entries to define the canonical 12 quarter columns.
    # Revenue is a duration metric present for every quarter, so its
    # period_end dates are the most reliable column anchors.
    rev_rows = [r for r in rows if r["metric_name"] == "Revenue"]
    rev_rows.sort(key=lambda r: r["period_end"], reverse=True)

    # Each column is identified by period_end; store (fy, fq, end).
    periods = [(r["fiscal_year"], r["fiscal_quarter"], r["period_end"])
               for r in rev_rows]

    metric_order = [
        "Revenue",
        "Operating Income",
        "Net Income",
        "Diluted EPS",
        "R&D Expense",
        "Operating Cash Flow",
        "Capital Expenditure",
        "Cash and Cash Equivalents",
    ]

    # Build lookup: (metric, period_end) -> row.
    # For balance-sheet (instant) items the period_end may not match exactly
    # because Qualcomm's fiscal calendar shifts by a day or two. We map each
    # instant entry to the nearest canonical period_end within 7 days.
    lookup = {}
    canonical_ends = [p[2] for p in periods]

    for r in rows:
        end = r["period_end"]
        if end in canonical_ends:
            lookup[(r["metric_name"], end)] = r
        else:
            from datetime import datetime
            r_date = datetime.strptime(end, "%Y-%m-%d").date()
            best_end = min(canonical_ends,
                           key=lambda ce: abs((datetime.strptime(ce, "%Y-%m-%d").date() - r_date).days))
            gap = abs((datetime.strptime(best_end, "%Y-%m-%d").date() - r_date).days)
            if gap <= 7:
                key = (r["metric_name"], best_end)
                if key not in lookup or r["filing_date"] > lookup[key]["filing_date"]:
                    lookup[key] = r

    # --- Write headers ---
    headers = ["Metric", "Unit"]
    for fy, fq, end in periods:
        headers.append(f"FY{fy} {fq}\n({end})")

    ws.append(headers)

    # --- Write data rows ---
    for metric in metric_order:
        row_data = [metric]
        first_match = next((lookup.get((metric, p[2])) for p in periods
                           if (metric, p[2]) in lookup), None)
        unit = first_match["unit"] if first_match else ""
        row_data.append(unit)

        for _, _, end in periods:
            entry = lookup.get((metric, end))
            if entry:
                row_data.append(float(entry["value"]))
            else:
                row_data.append("N/A")

        ws.append(row_data)

    # --- Formatting ---
    style_header_row(ws)

    # Column widths: Metric, Unit, then each quarter
    widths = {1: 28, 2: 12}
    for i in range(3, 3 + len(periods)):
        widths[i] = 18
    set_column_widths(ws, widths)

    # Apply number formats to data cells
    for row_idx in range(2, ws.max_row + 1):
        metric_name = ws.cell(row=row_idx, column=1).value
        for col_idx in range(3, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                if metric_name == "Diluted EPS":
                    cell.number_format = EPS_FORMAT
                else:
                    cell.number_format = USD_FORMAT

    # Wrap text in header row for multi-line quarter labels
    for cell in ws[1]:
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="bottom")

    # Add source info row at the bottom
    ws.append([])
    source_row = ws.max_row + 1
    ws.cell(row=source_row, column=1,
            value="Source: SEC EDGAR XBRL company-facts API")
    ws.cell(row=source_row, column=1).font = Font(italic=True)


# ---------------------------------------------------------------------------
# Tab 2: Missing Metrics
# ---------------------------------------------------------------------------

def build_missing_metrics(wb, missing_path: Path):
    ws = wb.create_sheet("Missing Metrics")

    with open(missing_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        missing_rows = list(reader)

    ws.append(["Metric", "Reason", "Suggestion"])

    for r in missing_rows:
        ws.append([r["metric"], r["reason"], r["suggestion"]])

    # Add metrics from CLAUDE.md that aren't extracted yet
    additional_gaps = [
        ["Gross Profit", "Not tagged as standalone XBRL fact by Qualcomm",
         "Derive from Revenue minus CostOfRevenue"],
        ["Total Debt", "Not yet extracted — requires identifying correct tags",
         "Sum of LongTermDebt + ShortTermBorrowings or similar"],
        ["Free Cash Flow", "Always a derived metric, never a raw XBRL tag",
         "Derive from Operating Cash Flow minus Capital Expenditure"],
        ["Gross Margin", "Ratio — not extracted in this pass",
         "Derive after Gross Profit is available"],
        ["Operating Margin", "Ratio — not extracted in this pass",
         "Derive from Operating Income / Revenue"],
        ["Free-Cash-Flow Margin", "Ratio — not extracted in this pass",
         "Derive from Free Cash Flow / Revenue"],
        ["R&D as % of Revenue", "Ratio — not extracted in this pass",
         "Derive from R&D Expense / Revenue"],
    ]

    existing = {r["metric"] for r in missing_rows}
    for row in additional_gaps:
        if row[0] not in existing:
            ws.append(row)

    style_header_row(ws)
    set_column_widths(ws, {1: 24, 2: 52, 3: 52})


# ---------------------------------------------------------------------------
# Tab 3: Data Dictionary
# ---------------------------------------------------------------------------

def build_data_dictionary(wb):
    ws = wb.create_sheet("Data Dictionary")

    ws.append(["Field", "Description"])

    dictionary = [
        ["company", "Legal entity name (Qualcomm Incorporated)"],
        ["fiscal_year", "Qualcomm fiscal year (ends late September)"],
        ["fiscal_quarter", "Q1–Q3 from 10-Q filings; FY from 10-K"],
        ["period_start", "First day of the reporting period (blank for balance-sheet items)"],
        ["period_end", "Last day of the reporting period"],
        ["filing_date", "Date the SEC filing was submitted"],
        ["sec_form_type", "10-Q (quarterly) or 10-K (annual)"],
        ["metric_name", "Financial metric label (see Metric Reference below)"],
        ["value", "Reported value in the unit shown"],
        ["unit", "USD or USD/shares"],
        ["source_reference", "URL to the specific SEC filing on EDGAR"],
    ]

    for row in dictionary:
        ws.append(row)

    ws.append([])
    ws.append(["Metric Reference", "XBRL Tag Used"])

    metric_tags = [
        ["Revenue", "us-gaap:Revenues"],
        ["Operating Income", "us-gaap:OperatingIncomeLoss"],
        ["Net Income", "us-gaap:NetIncomeLoss"],
        ["Diluted EPS", "us-gaap:EarningsPerShareDiluted"],
        ["Cash and Cash Equivalents", "us-gaap:CashAndCashEquivalentsAtCarryingValue"],
        ["Operating Cash Flow", "us-gaap:NetCashProvidedByUsedInOperatingActivities"],
        ["Capital Expenditure", "us-gaap:PaymentsToAcquireProductiveAssets"],
        ["R&D Expense", "us-gaap:ResearchAndDevelopmentExpense"],
    ]

    for row in metric_tags:
        ws.append(row)

    style_header_row(ws)
    set_column_widths(ws, {1: 30, 2: 58})

    # Also bold the "Metric Reference" sub-header
    for row in ws.iter_rows():
        if row[0].value == "Metric Reference":
            row[0].font = HEADER_FONT
            row[1].font = HEADER_FONT
            break


# ---------------------------------------------------------------------------
# Tab 4: Manual Checks
# ---------------------------------------------------------------------------

def build_manual_checks(wb):
    ws = wb.create_sheet("Manual Checks")

    ws.append(["Item", "Detail", "Action Required"])

    checks = [
        [
            "No Q4 standalone data",
            "SEC XBRL reports Q4 results inside the annual 10-K as a "
            "full-year figure. There is no standalone Q4 tag.",
            "Derive Q4 = FY total minus (Q1 + Q2 + Q3) when annual data is added",
        ],
        [
            "Gross Profit not tagged",
            "Qualcomm does not file a GrossProfit XBRL fact. "
            "CostOfRevenue is available.",
            "Derive as Revenue minus CostOfRevenue and verify against 10-K",
        ],
        [
            "Capital Expenditure tag",
            "Script uses PaymentsToAcquireProductiveAssets because Qualcomm "
            "does not file PaymentsToAcquirePropertyPlantAndEquipment. "
            "This broader tag may include non-PP&E productive assets.",
            "Cross-check a recent quarter against the 10-Q cash flow statement",
        ],
        [
            "Fiscal-year label assignment",
            "The same calendar quarter can appear under two fiscal-year labels "
            "across filings. The script keeps the most recently filed version, "
            "so the FY/FQ label comes from that filing.",
            "Verify FY/FQ labels match Qualcomm's investor-relations calendar",
        ],
        [
            "Balance-sheet items are point-in-time",
            "Cash and Cash Equivalents has no period_start — it is a snapshot "
            "at period_end, not a flow over a range.",
            "No action needed; this is expected accounting treatment",
        ],
    ]

    for row in checks:
        ws.append(row)

    style_header_row(ws)
    set_column_widths(ws, {1: 30, 2: 58, 3: 52})

    # Wrap text for readability
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    csv_path = find_latest_file(PROCESSED_DIR, "qualcomm_financials_*.csv")
    missing_path = find_latest_file(MANUAL_CHECKS_DIR, "qualcomm_missing_metrics_*.csv")

    print(f"Reading: {csv_path.name}")
    print(f"Reading: {missing_path.name}")

    with open(csv_path, "r", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    wb = openpyxl.Workbook()

    build_quarterly_financials(wb, rows)
    build_missing_metrics(wb, missing_path)
    build_data_dictionary(wb)
    build_manual_checks(wb)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / "qualcomm_financial_history.xlsx"
    wb.save(out_path)

    print(f"\nWorkbook saved to: {out_path}")
    print(f"Tabs: {wb.sheetnames}")
    print(f"Quarterly Financials: {wb['Quarterly Financials'].max_row - 1} metrics x "
          f"{wb['Quarterly Financials'].max_column - 2} quarters")


if __name__ == "__main__":
    main()
